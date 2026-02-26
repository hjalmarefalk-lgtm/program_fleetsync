"""Phase 4 post-write validator: reopen workbook and validate invariants (no UI)."""

from __future__ import annotations

from pathlib import Path
import re
from typing import List
from collections import Counter

from openpyxl import load_workbook
from zipfile import BadZipFile

from .models import ValidationIssue, ValidationReport
from .validation import compute_valid, issue

_INVALID_SHEET_CHARS = re.compile(r"[:\\/\?\*\[\]]")


def _header_values(ws) -> list[str]:
    try:
        row = next(ws.iter_rows(min_row=1, max_row=1))
    except StopIteration:
        return []
    return [str(cell.value) if cell.value is not None else "" for cell in row]


def validate_written_workbook(path: str | Path) -> ValidationReport:
    """Validate exported workbook invariants and return a ValidationReport."""
    warnings: List[ValidationIssue] = []
    fatals: List[ValidationIssue] = []

    workbook_path = Path(path)
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
    except (BadZipFile, OSError) as exc:
        fatals.append(
            issue(
                "POSTWRITE_BADZIP",
                "FATAL",
                f"Workbook is not a valid .xlsx file: {workbook_path}",
            )
        )
        return ValidationReport(valid=compute_valid(warnings, fatals), warnings=warnings, fatals=fatals)
    try:
        sheet_names = wb.sheetnames

        has_summary = "Sammanfattning" in sheet_names
        has_total = "Total" in sheet_names

        if len(sheet_names) != len(set(sheet_names)):
            fatals.append(issue("POST_001", "FATAL", "Duplicate sheet names detected"))

        for name in sheet_names:
            if len(name) > 31:
                fatals.append(issue("POST_002", "FATAL", f"Sheet name exceeds 31 chars: {name}"))
            if _INVALID_SHEET_CHARS.search(name):
                fatals.append(issue("POST_003", "FATAL", f"Sheet name contains invalid chars: {name}"))

        # P9.9 contract: summary sheet is optional, but if present must be first.
        if has_summary and sheet_names[0] != "Sammanfattning":
            fatals.append(issue("POST_005", "FATAL", "Summary sheet must be first"))

        data_sheet_names = [n for n in sheet_names if n != "Sammanfattning"]

        # P9.9 contract: if Total sheet is absent, there must be exactly one data sheet.
        if not has_total:
            if len(data_sheet_names) != 1:
                fatals.append(issue("POST_006", "FATAL", "Total sheet missing (split-mode output)"))
            else:
                main_headers = _header_values(wb[data_sheet_names[0]])
                if not main_headers:
                    warnings.append(issue("POST_007", "WARNING", "Data sheet header row is empty"))
        else:
            total_headers: list[str] = _header_values(wb["Total"])
            if not total_headers:
                warnings.append(issue("POST_007", "WARNING", "Total sheet header row is empty"))

            split_sheet_names = [name for name in sheet_names if name not in ("Total", "Sammanfattning")]
            for name in sheet_names:
                if name == "Sammanfattning":
                    continue
                headers = _header_values(wb[name])
                if total_headers and headers and headers != total_headers:
                    warnings.append(
                        issue(
                            "POST_008",
                            "WARNING",
                            f"Column order mismatch vs Total in sheet: {name}",
                        )
                    )

            if split_sheet_names and total_headers:
                union_counter: Counter[tuple] = Counter()
                header_mismatch = False
                for name in split_sheet_names:
                    headers = _header_values(wb[name])
                    if headers and headers != total_headers:
                        warnings.append(
                            issue(
                                "POST_010",
                                "WARNING",
                                "Skipping Total union check due to header mismatch",
                            )
                        )
                        header_mismatch = True
                        break
                    for row in wb[name].iter_rows(min_row=2, values_only=True):
                        union_counter[tuple(row)] += 1

                if not header_mismatch:
                    total_counter: Counter[tuple] = Counter()
                    for row in wb["Total"].iter_rows(min_row=2, values_only=True):
                        total_counter[tuple(row)] += 1
                    for row_key, count in total_counter.items():
                        if count > union_counter.get(row_key, 0):
                            fatals.append(
                                issue(
                                    "TOTAL_EXTRAS",
                                    "FATAL",
                                    "Total contains rows not present in union of split sheets",
                                )
                            )
                            break
    finally:
        wb.close()

    return ValidationReport(valid=compute_valid(warnings, fatals), warnings=warnings, fatals=fatals)
