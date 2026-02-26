"""File probing service.

X-Ray v1 requirements:
- Read only a sample of rows (default nrows=300).
- Never display or log cell values.

This module returns headers + in-memory samples for classification only.
"""

from __future__ import annotations

import csv
import os
from dataclasses import dataclass
from typing import Dict, Iterable, List, Literal, Mapping, Sequence

from .probe_scan_constants import HEADER_SCAN_M, HEADER_SCAN_N, PREVIEW_SCAN_M, PREVIEW_SCAN_N


FileKind = Literal["xlsx", "csv", "unsupported"]

# LOCK: central scan window constants live in services/probe_scan_constants.py


def detect_file_kind(input_path: str) -> FileKind:
    ext = os.path.splitext(input_path or "")[1].lower()
    if ext == ".csv":
        return "csv"
    if ext in (".xlsx", ".xlsm"):
        return "xlsx"
    return "unsupported"


@dataclass(frozen=True, slots=True)
class ProbeResult:
    input_path: str
    file_kind: FileKind
    workbook_sheets: tuple[str, ...]
    selected_sheet_name: str | None
    ws_max_row: int
    ws_max_col: int
    # P6.1: bounded scan window diagnostics (used by sheet preview + future header detection).
    sample_rows: int
    sample_cols: int
    # P6.2: header detection diagnostics
    detected_header_row: int | None
    best_row_nonempty_count: int
    header_nonempty_count: int
    effective_width: int
    # P6.7: structural cap counter (header scan only; values are never retained)
    header_scan_cells_scanned: int
    headers: tuple[str, ...]
    samples_by_header: Mapping[str, tuple[object, ...]]
    sampled_rows: int


def probe_file(
    *,
    input_path: str,
    sheet_name: str | None = None,
    nrows: int = 300,
    k_values: int = 50,
) -> ProbeResult:
    """Probe a file and return headers and sampled values per column."""

    if not input_path:
        raise ValueError("No file selected")
    if not os.path.exists(input_path):
        raise ValueError("File not found")

    kind = detect_file_kind(input_path)
    if kind == "csv":
        headers, samples, sampled_rows = _probe_csv(input_path, nrows=nrows, k_values=k_values)
        return ProbeResult(
            input_path=input_path,
            file_kind=kind,
            workbook_sheets=(),
            selected_sheet_name=None,
            ws_max_row=0,
            ws_max_col=0,
            sample_rows=0,
            sample_cols=0,
            detected_header_row=None,
            best_row_nonempty_count=0,
            header_nonempty_count=0,
            effective_width=len(headers),
            header_scan_cells_scanned=0,
            headers=tuple(headers),
            samples_by_header={h: tuple(vs) for h, vs in samples.items()},
            sampled_rows=int(sampled_rows),
        )

    if kind == "xlsx":
        (
            sheets,
            selected,
            headers,
            samples,
            sampled_rows,
            ws_max_row,
            ws_max_col,
            detected_header_row,
            best_row_nonempty_count,
            header_nonempty_count,
            effective_width,
            header_scan_cells_scanned,
        ) = _probe_xlsx(
            input_path,
            sheet_name=sheet_name,
            nrows=nrows,
            k_values=k_values,
        )
        return ProbeResult(
            input_path=input_path,
            file_kind=kind,
            workbook_sheets=tuple(sheets),
            selected_sheet_name=selected,
            ws_max_row=int(ws_max_row),
            ws_max_col=int(ws_max_col),
            sample_rows=HEADER_SCAN_N,
            sample_cols=HEADER_SCAN_M,
            detected_header_row=detected_header_row,
            best_row_nonempty_count=int(best_row_nonempty_count),
            header_nonempty_count=int(header_nonempty_count),
            effective_width=int(effective_width),
            header_scan_cells_scanned=int(header_scan_cells_scanned),
            headers=tuple(headers),
            samples_by_header={h: tuple(vs) for h, vs in samples.items()},
            sampled_rows=int(sampled_rows),
        )

    raise ValueError("Unsupported file type")


def _probe_csv(
    input_path: str,
    *,
    nrows: int,
    k_values: int,
) -> tuple[list[str], Dict[str, list[object]], int]:
    # Use Python's csv reader to avoid additional dependencies.
    with open(input_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        try:
            header_row = next(reader)
        except StopIteration:
            raise ValueError("Empty CSV")

        headers = _normalize_headers(header_row)
        samples: Dict[str, list[object]] = {h: [] for h in headers}

        sampled_rows = 0

        for row_idx, row in enumerate(reader, start=1):
            if row_idx > nrows:
                break
            sampled_rows += 1
            # Pad / truncate rows to header width deterministically.
            row_cells = list(row[: len(headers)])
            if len(row_cells) < len(headers):
                row_cells.extend([""] * (len(headers) - len(row_cells)))

            for header, cell in zip(headers, row_cells, strict=True):
                if len(samples[header]) >= k_values:
                    continue
                if cell is None:
                    continue
                s = str(cell).strip()
                if s == "":
                    continue
                samples[header].append(s)

            if _all_samples_full(samples, k_values=k_values):
                break

        return headers, samples, sampled_rows


def _probe_xlsx(
    input_path: str,
    *,
    sheet_name: str | None,
    nrows: int,
    k_values: int,
) -> tuple[list[str], str | None, list[str], Dict[str, list[object]], int, int, int, int | None, int, int, int, int]:
    try:
        from openpyxl import load_workbook
    except Exception as e:  # pragma: no cover
        raise ValueError("Missing openpyxl") from e

    wb = load_workbook(filename=input_path, read_only=True, data_only=True)
    try:
        sheetnames = list(getattr(wb, "sheetnames", []) or [])
        if not sheetnames:
            raise ValueError("No worksheets")

        requested = (sheet_name or "").strip()
        if requested and requested in sheetnames:
            selected = requested
        else:
            selected = _first_previewable_sheet(wb, sheetnames)

        if not selected:
            # No previewable sheets: return a safe empty probe.
            return sheetnames, None, [], {}, 0, 0, 0, None, 0, 0, 0, 0

        (
            headers,
            samples,
            sampled_rows,
            ws_max_row,
            ws_max_col,
            detected_header_row,
            best_row_nonempty_count,
            header_nonempty_count,
            effective_width,
            header_scan_cells_scanned,
        ) = _probe_xlsx_sheet(
            wb,
            selected,
            nrows=nrows,
            k_values=k_values,
        )
        return (
            sheetnames,
            selected,
            headers,
            samples,
            sampled_rows,
            ws_max_row,
            ws_max_col,
            detected_header_row,
            best_row_nonempty_count,
            header_nonempty_count,
            effective_width,
            header_scan_cells_scanned,
        )
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _first_previewable_sheet(wb, sheetnames: Sequence[str]) -> str | None:
    for name in sheetnames:
        try:
            ws = wb[name]
        except Exception:
            continue
        if _is_previewable_sheet(ws):
            return name
    return None


def _is_previewable_sheet(ws) -> bool:
    # LOCK: cheap deterministic scan, N=50, M=50.
    try:
        max_row = int(getattr(ws, "max_row", 0) or 0)
        max_col = int(getattr(ws, "max_column", 0) or 0)
    except Exception:
        max_row = 0
        max_col = 0

    if max_row < 1 or max_col < 1:
        # Still scan a bounded window; openpyxl metadata may be 1 even for empties.
        max_row = PREVIEW_SCAN_N
        max_col = PREVIEW_SCAN_M

    scan_rows = min(PREVIEW_SCAN_N, max_row)
    scan_cols = min(PREVIEW_SCAN_M, max_col)
    if scan_rows < 1 or scan_cols < 1:
        return False

    try:
        for row in ws.iter_rows(
            min_row=1,
            max_row=scan_rows,
            min_col=1,
            max_col=scan_cols,
            values_only=True,
        ):
            for cell in row:
                if cell is None:
                    continue
                if isinstance(cell, str) and cell.strip() == "":
                    continue
                return True
        return False
    except Exception:
        return False


def _probe_xlsx_sheet(
    wb,
    sheet_name: str,
    *,
    nrows: int,
    k_values: int,
) -> tuple[list[str], Dict[str, list[object]], int, int, int, int | None, int, int, int, int]:
    try:
        ws = wb[sheet_name]
    except Exception:
        # If the requested sheet cannot be accessed, treat as empty.
        return [], {}, 0, 0, 0, None, 0, 0, 0

    try:
        ws_max_row = int(getattr(ws, "max_row", 0) or 0)
        ws_max_col = int(getattr(ws, "max_column", 0) or 0)
    except Exception:
        ws_max_row = 0
        ws_max_col = 0

    def _is_nonempty(cell: object) -> bool:
        if cell is None:
            return False
        s = str(cell)
        return s.strip() != ""

    # P6.2: deterministically pick densest row within a bounded window.
    best_row_index: int | None = None
    best_count = -1
    max_data_last = 0
    header_scan_cells_scanned = 0

    try:
        for r_idx, row in enumerate(
            ws.iter_rows(
                min_row=1,
                max_row=HEADER_SCAN_N,
                min_col=1,
                max_col=HEADER_SCAN_M,
                values_only=True,
            ),
            start=1,
        ):
            header_scan_cells_scanned += len(row)
            nonempty = 0
            row_last = 0
            for c_idx, cell in enumerate(row, start=1):
                if not _is_nonempty(cell):
                    continue
                nonempty += 1
                row_last = c_idx

            if row_last > max_data_last:
                max_data_last = row_last

            if nonempty > best_count or (nonempty == best_count and (best_row_index is None or r_idx < best_row_index)):
                best_count = nonempty
                best_row_index = r_idx
    except Exception:
        return [], {}, 0, ws_max_row, ws_max_col, None, 0, 0, 0, int(header_scan_cells_scanned)

    best_row_nonempty_count = max(0, int(best_count))

    if best_row_index is None:
        return [], {}, 0, ws_max_row, ws_max_col, None, 0, 0, max(0, int(max_data_last)), int(header_scan_cells_scanned)

    if best_row_nonempty_count < 2 or best_row_index is None:
        # Header floor (P6.2): no detected header row.
        detected_header_row = None
        header_last = 0
        effective_width = max(int(header_last), int(max_data_last))
        if effective_width <= 0:
            return [], {}, 0, ws_max_row, ws_max_col, None, best_row_nonempty_count, 0, 0, int(header_scan_cells_scanned)

        headers = _normalize_headers([None] * int(effective_width))
        samples: Dict[str, list[object]] = {h: [] for h in headers}

        sampled_rows = 0
        max_row = 1 + max(0, nrows)
        for row in ws.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=HEADER_SCAN_M,
            values_only=True,
        ):
            sampled_rows += 1
            row_cells = list(row[: len(headers)])
            if len(row_cells) < len(headers):
                row_cells.extend([None] * (len(headers) - len(row_cells)))

            for header, cell in zip(headers, row_cells, strict=True):
                if len(samples[header]) >= k_values:
                    continue
                if not _is_nonempty(cell):
                    continue
                samples[header].append(cell)

            if _all_samples_full(samples, k_values=k_values):
                break

        return (
            headers,
            samples,
            sampled_rows,
            ws_max_row,
            ws_max_col,
            detected_header_row,
            best_row_nonempty_count,
            0,
            int(effective_width),
            int(header_scan_cells_scanned),
        )

    detected_header_row = int(best_row_index)

    # Read header labels from the detected row, bounded by M.
    header_cells = next(
        ws.iter_rows(
            min_row=detected_header_row,
            max_row=detected_header_row,
            min_col=1,
            max_col=HEADER_SCAN_M,
            values_only=True,
        ),
        None,
    )
    if header_cells is None:
        return (
            [],
            {},
            0,
            ws_max_row,
            ws_max_col,
            None,
            best_row_nonempty_count,
            0,
            max(0, int(max_data_last)),
            int(header_scan_cells_scanned),
        )

    header_list = list(header_cells)
    last_nonempty_col = 0
    for idx, cell in enumerate(header_list, start=1):
        if _is_nonempty(cell):
            last_nonempty_col = idx

    header_last = int(last_nonempty_col) if last_nonempty_col >= 1 else 0
    effective_width = max(int(header_last), int(max_data_last))

    if effective_width < 2:
        return (
            [],
            {},
            0,
            ws_max_row,
            ws_max_col,
            None,
            best_row_nonempty_count,
            0,
            int(effective_width),
            int(header_scan_cells_scanned),
        )

    header_slice = header_list[: int(effective_width)]
    header_nonempty_count = sum(1 for c in header_slice if _is_nonempty(c))
    headers = _normalize_headers(header_slice)
    samples: Dict[str, list[object]] = {h: [] for h in headers}

    sampled_rows = 0
    max_row = 1 + max(0, nrows)
    for row in ws.iter_rows(
        min_row=detected_header_row + 1,
        max_row=max_row,
        min_col=1,
        max_col=HEADER_SCAN_M,
        values_only=True,
    ):
        sampled_rows += 1
        row_cells = list(row[: len(headers)])
        if len(row_cells) < len(headers):
            row_cells.extend([None] * (len(headers) - len(row_cells)))

        for header, cell in zip(headers, row_cells, strict=True):
            if len(samples[header]) >= k_values:
                continue
            if not _is_nonempty(cell):
                continue
            samples[header].append(cell)

        if _all_samples_full(samples, k_values=k_values):
            break

    return (
        headers,
        samples,
        sampled_rows,
        ws_max_row,
        ws_max_col,
        detected_header_row,
        best_row_nonempty_count,
        header_nonempty_count,
        int(effective_width),
        int(header_scan_cells_scanned),
    )


def _normalize_headers(raw_headers: Sequence[object]) -> list[str]:
    headers: list[str] = []
    for idx, h in enumerate(raw_headers, start=1):
        if h is None:
            name = f"Column {idx}"
        else:
            name = str(h).strip()
            if name == "":
                name = f"Column {idx}"
        headers.append(name)
    return headers


def _all_samples_full(samples: Mapping[str, Sequence[object]], *, k_values: int) -> bool:
    return all(len(v) >= k_values for v in samples.values())
