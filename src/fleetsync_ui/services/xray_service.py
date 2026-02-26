"""X-Ray probe wrapper (canonical entry point).

LOCK (P6.4): UI must call this wrapper (directly or via worker).

Contract:
- file_key = (abs_path, mtime_ns, size_bytes)
- probe_key = (abs_path, mtime_ns, size_bytes, sheet_name)
- XLSX sheet ordering is preserved as workbook order.
- Never returns sampled cell values.
"""

from __future__ import annotations

from typing import Optional

from .dto import ProbeXRayPayload
from .file_probe import detect_file_kind, probe_file
from .probe_cache import PROBE_CACHE, ProbeCache, compute_file_key, compute_probe_key
from .xray_classifier import classify_columns


_SHEETS_SUFFIX = "__SHEETS__"
_AUTO_SHEET = "__AUTO__"

SHEETS_CACHE: ProbeCache[tuple[str, ...]] = ProbeCache(max_entries=5)


def _get_workbook_sheetnames(*, input_path: str) -> list[str]:
    """Read workbook sheetnames in workbook order.

    Kept separate so cache-hit paths can still populate the sheet selector without
    recomputing the full probe.
    """

    try:
        from openpyxl import load_workbook
    except Exception as e:  # pragma: no cover
        raise ValueError("Missing openpyxl") from e

    wb = load_workbook(filename=input_path, read_only=True, data_only=True)
    try:
        return list(getattr(wb, "sheetnames", []) or [])
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _get_cached_sheetnames(*, input_path: str) -> tuple[str, ...] | None:
    key = compute_probe_key(input_path, sheet_name=_SHEETS_SUFFIX)
    if key is None:
        return None
    return SHEETS_CACHE.get(key)


def _set_cached_sheetnames(*, input_path: str, sheets: list[str]) -> None:
    key = compute_probe_key(input_path, sheet_name=_SHEETS_SUFFIX)
    if key is None:
        return
    SHEETS_CACHE.set(key, tuple(sheets))


def probe_xray(*, input_path: str, sheet_name: Optional[str] = None) -> ProbeXRayPayload:
    """Canonical probe wrapper used by UI.

    Returns a fully UI-safe payload. Uses in-memory cache keyed by sheet.
    """

    kind = detect_file_kind(input_path)
    if kind == "unsupported":
        raise ValueError("Unsupported file type")

    file_key = compute_file_key(input_path)

    if kind == "csv":
        probe_key = compute_probe_key(input_path, sheet_name="__CSV__")
        if probe_key is not None:
            cached = PROBE_CACHE.get(probe_key)
            if cached is not None:
                return ProbeXRayPayload(
                    input_path=input_path,
                    file_kind="csv",
                    file_key=file_key,
                    workbook_sheets=[],
                    preview_sheet_name=None,
                    xray_status="ready",
                    xray_result=cached,
                    ui_notice=None,
                )

        probe = probe_file(input_path=input_path)

        best_count = int(getattr(probe, "best_row_nonempty_count", 0) or 0)
        effective_width = int(getattr(probe, "effective_width", 0) or 0)
        confidence_raw = min(1.0, best_count / 10.0) * min(1.0, effective_width / 20.0)
        confidence_display = f"{confidence_raw:.2f}"
        result = classify_columns(
            input_path=probe.input_path,
            sheet_name="__CSV__",
            sheet_index=0,
            ws_max_row=int(getattr(probe, "ws_max_row", 0) or 0),
            ws_max_col=int(getattr(probe, "ws_max_col", 0) or 0),
            sample_rows=int(getattr(probe, "sample_rows", 0) or 0),
            sample_cols=int(getattr(probe, "sample_cols", 0) or 0),
            detected_header_row=getattr(probe, "detected_header_row", None),
            best_row_nonempty_count=best_count,
            header_nonempty_count=int(getattr(probe, "header_nonempty_count", 0) or 0),
            effective_width=effective_width,
            header_scan_cells_scanned=int(getattr(probe, "header_scan_cells_scanned", 0) or 0),
            confidence_raw=confidence_raw,
            confidence_display=confidence_display,
            headers=probe.headers,
            samples_by_header=probe.samples_by_header,
            sampled_rows=probe.sampled_rows,
        )

        if probe_key is not None:
            PROBE_CACHE.set(probe_key, result)

        return ProbeXRayPayload(
            input_path=probe.input_path,
            file_kind="csv",
            file_key=file_key,
            workbook_sheets=[],
            preview_sheet_name=None,
            xray_status="ready",
            xray_result=result,
            ui_notice=None,
        )

    # XLSX
    requested = (sheet_name or "").strip()
    cache_sheet = requested if requested else _AUTO_SHEET

    probe_key = compute_probe_key(input_path, sheet_name=cache_sheet)
    if probe_key is not None:
        cached = PROBE_CACHE.get(probe_key)
        if cached is not None:
            sheets = _get_cached_sheetnames(input_path=input_path)
            if sheets is None:
                sheets_list = _get_workbook_sheetnames(input_path=input_path)
                _set_cached_sheetnames(input_path=input_path, sheets=sheets_list)
                sheets = tuple(sheets_list)

            return ProbeXRayPayload(
                input_path=input_path,
                file_kind="xlsx",
                file_key=file_key,
                workbook_sheets=list(sheets or ()),
                preview_sheet_name=cached.sheet_name or None,
                xray_status="ready",
                xray_result=cached,
                ui_notice=None,
            )

    probe = probe_file(input_path=input_path, sheet_name=(requested if requested else None))

    # Workbook loaded, but no previewable sheets.
    if probe.file_kind == "xlsx" and probe.selected_sheet_name is None:
        return ProbeXRayPayload(
            input_path=probe.input_path,
            file_kind="xlsx",
            file_key=file_key,
            workbook_sheets=list(probe.workbook_sheets),
            preview_sheet_name=None,
            xray_status="empty",
            xray_result=None,
            ui_notice="No previewable sheets (all empty).",
        )

    selected = probe.selected_sheet_name or ""
    sheet_index = (
        list(probe.workbook_sheets).index(probe.selected_sheet_name)
        if (probe.selected_sheet_name in probe.workbook_sheets)
        else -1
    )

    best_count = int(getattr(probe, "best_row_nonempty_count", 0) or 0)
    effective_width = int(getattr(probe, "effective_width", 0) or 0)
    confidence_raw = min(1.0, best_count / 10.0) * min(1.0, effective_width / 20.0)
    confidence_display = f"{confidence_raw:.2f}"

    result = classify_columns(
        input_path=probe.input_path,
        sheet_name=selected,
        sheet_index=sheet_index,
        ws_max_row=int(getattr(probe, "ws_max_row", 0) or 0),
        ws_max_col=int(getattr(probe, "ws_max_col", 0) or 0),
        sample_rows=int(getattr(probe, "sample_rows", 0) or 0),
        sample_cols=int(getattr(probe, "sample_cols", 0) or 0),
        detected_header_row=getattr(probe, "detected_header_row", None),
        best_row_nonempty_count=best_count,
        header_nonempty_count=int(getattr(probe, "header_nonempty_count", 0) or 0),
        effective_width=effective_width,
        header_scan_cells_scanned=int(getattr(probe, "header_scan_cells_scanned", 0) or 0),
        confidence_raw=confidence_raw,
        confidence_display=confidence_display,
        headers=probe.headers,
        samples_by_header=probe.samples_by_header,
        sampled_rows=probe.sampled_rows,
    )

    # Cache under the actual selected sheet if explicit and valid; else auto.
    if requested and selected == requested:
        key = compute_probe_key(input_path, sheet_name=selected)
    else:
        key = compute_probe_key(input_path, sheet_name=_AUTO_SHEET)

    if key is not None:
        PROBE_CACHE.set(key, result)

    _set_cached_sheetnames(input_path=input_path, sheets=list(probe.workbook_sheets))

    return ProbeXRayPayload(
        input_path=probe.input_path,
        file_kind="xlsx",
        file_key=file_key,
        workbook_sheets=list(probe.workbook_sheets),
        preview_sheet_name=probe.selected_sheet_name,
        xray_status="ready",
        xray_result=result,
        ui_notice=None,
    )
